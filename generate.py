import argparse
import os

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from datetime import datetime
import itertools


def maxstr(l):
    return len(max(l, key=len))


def generate_code(
        class_name,
        states_dict,
        events,
        state_table,
        connectors,
        date=None,
        else_event="ELSE"
):
    """
    states_enum: state names (will be converted to upper case)
    events_enum: will be converted to a list of [EVT_(UPPER_CASE_EVENTS), ELSE]

    """
    # states_name, states_action = zip(*[(k.upper(), v) for k, v in states_dict.items()])

    states_name = states_dict.keys()
    states_enum = ", ".join(states_name)

    # store actions used in enums like ENT_STATE1
    enters = [x["enter"] for x in states_dict.values() if x["enter"] != "-1"]
    loops = [x["loop"] for x in states_dict.values() if x["loop"] != "-1"]
    exits = [x["exit"] for x in states_dict.values() if x["exit"] != "-1"]
    actions_enum = ", ".join(enters + loops + exits)

    # store actions used in state table. No line will be dropped.
    enters_table_col = [x["enter"] for x in states_dict.values()]
    loops_table_col = [x["loop"] for x in states_dict.values()]
    exits_table_col = [x["exit"] for x in states_dict.values()]

    # Connectors
    if connectors:
        slot_index = 0
        connectors_enum = []
        connector_shortcuts_def = ""
        for name, connector_dict in connectors.items():
            auto_store = connector_dict.get("auto_store", 0)
            enum_name = connector_dict.get("enum_name", 0) or f"ON_{name.upper()}"
            function_name = connector_dict.get("function_name", 0) or f"on{name.lower().capitalize()}"
            broadcast = connector_dict.get("broadcast", 0)

            connectors_enum.append(f"{enum_name}={slot_index}")
            slot = connector_dict["slot"] if "slot" in connector_dict else 1
            slot_index += slot


            # Generate function definitions.
            if auto_store:
                connector_shortcuts_def += f"""
  {class_name}& {function_name}( Machine& machine, int event ) {{ onPush( connectors, {enum_name}, -1, {slot}, 0, machine, event ); return *this; }}
  {class_name}& {function_name}( atm_cb_push_t callback, int idx ) {{ onPush( connectors, {enum_name}, -1, {slot}, 0, callback, idx ); return *this; }}
"""
            elif broadcast:
                connector_shortcuts_def += f"""
  {class_name}& {function_name}( Machine& machine, int event ) {{ onPush( connectors, {enum_name}, 0, {slot}, 1, machine, event ); return *this; }}
  {class_name}& {function_name}( atm_cb_push_t callback, int idx ) {{ onPush( connectors, {enum_name}, 0, {slot}, 1, callback, idx ); return *this; }}
  {class_name}& {function_name}( int sub, Machine& machine, int event ) {{ onPush( connectors, {enum_name}, sub, {slot}, 0, machine, event ); return *this; }}
  {class_name}& {function_name}( int sub, atm_cb_push_t callback, int idx ) {{ onPush( connectors, {enum_name}, sub, {slot}, 0, callback, idx ); return *this; }}
"""
            else:
                connector_shortcuts_def += f"""
  {class_name}& {function_name}( Machine& machine, int event ) {{ onPush( connectors, {enum_name}, 0, {slot}, 1, machine, event ); return *this; }}
  {class_name}& {function_name}( atm_cb_push_t callback, int idx ) {{ onPush( connectors, {enum_name}, 0, {slot}, 1, callback, idx ); return *this; }}
"""

        connectors_enum.append(f"CONN_MAX={slot_index}")
        connectors_code = f"enum {{ {', '.join(connectors_enum)} }}; // CONNECTORS" + "\n" + "  atm_connector connectors[CONN_MAX]; // CONNECTORS"
    else:
        connectors_code = ""
        connector_shortcuts_def = ""

    # Events
    events_enum = ", ".join(events.values())
    events_name = events.values()
    trigger_shortcuts_def = ""
    for event, en in zip(events,events_name):
        event = event.lower()
        if event.upper() == else_event:
            continue
        trigger_shortcuts_def += f"  {class_name}& {event}() {{ trigger({en}); return *this; }}\n"

    # src part
    state_table_by_evt_column = list(zip(*state_table))

    col_max_len = []

    # first row
    states_name_max_len = maxstr(states_name)
    enters_max_len = maxstr(enters_table_col + ["ON_ENTER"])
    loops_max_len = maxstr(loops_table_col + ["ON_LOOP"])
    exits_max_len = maxstr(exits_table_col + ["ON_EXIT"])
    col_max_len.extend([states_name_max_len, enters_max_len, loops_max_len, exits_max_len])
    spaces = " " * 6
    state_table_str = f"{spaces}/* {'':>{states_name_max_len + 3}}  {'ON_ENTER':>{enters_max_len}}  {'ON_LOOP':>{loops_max_len}}  {'ON_EXIT':>{exits_max_len}}  "
    # appending events
    for evt_col, en in zip(state_table_by_evt_column, events_name):
        max_len = maxstr(evt_col + (en,))
        state_table_str += f"{en:>{max_len}}  "
        col_max_len.append(max_len)

    state_table_str += "*/\n"

    # cache precomputed widths
    # rest row
    for sn, en, lo, ex, st in zip(states_name, enters_table_col, loops_table_col, exits_table_col, state_table):
        state_table_str += f"{spaces}/* {sn:>{col_max_len[0]}} */  {en:>{col_max_len[1]}}, {lo:>{col_max_len[2]}}, {ex:>{col_max_len[3]}}, "
        for st_col, col_len in zip(st, col_max_len[4:]):
            assert st_col == "" or st_col in states_name
            st_col = st_col or "-1"
            state_table_str += f"{st_col:>{col_len}}, "
        state_table_str += "\n"
    state_table_str = state_table_str.rstrip()

    trace_string = "\\0".join([class_name.upper(), *events_name, *states_name])
    header = f"""#pragma once
/* 
 * Generated by automaton_codegen on {date or datetime.now().strftime("%x %X")}
 * Subclass this class and implement the features to use.
 */

#include <Automaton.h>

class {class_name}: public Machine {{
 public:
  enum {{ {states_enum} }}; // STATES
  enum {{ {events_enum} }}; // EVENTS
  {class_name}() : Machine() {{}}
  virtual {class_name}& begin() {{
    // clang-format off
    const static state_t state_table[] PROGMEM = {{
{state_table_str}
    }};
    // clang-format on
    Machine::begin(state_table, ELSE);
    return *this;
  }}
  {class_name}& trace( Stream & stream ) {{
    Machine::setTrace(&stream, atm_serial_debug::trace, "{trace_string}");
    return *this;
  }}
  {class_name}& trigger( int event ) {{
    Machine::trigger(event);
    return *this;
  }};
  // connector shortcuts
{connector_shortcuts_def}
  // trigger shortcuts
{trigger_shortcuts_def}

 private:
  enum {{ {actions_enum} }}; // ACTIONS
  {connectors_code}
  virtual int event( int id ) = 0;
  virtual void action( int id ) = 0;
}};

"""
    return header


def read_xls(xls_file):
    wb = load_workbook(xls_file)
    return wb


def parse_workbook(wb: Workbook):
    states_ws: Worksheet = wb["states"]
    states = {}
    for r in list(states_ws.rows)[1:]:
        state_cell: Cell = r[0]
        state_name = state_cell.value.upper()
        enter, loop, exit = r[1:4]

        states[state_cell.value] = {
            "enter": f"ENT_{state_name}" if enter.value else "-1",
            "loop": f"LP_{state_name}" if loop.value == 1 else ("ATM_SLEEP" if loop.value == "sleep" else "-1"),
            "exit": f"EXT_{state_name}" if exit.value else "-1"
        }

    events = {}
    for c in list(states_ws.rows)[0][5:]:
        evt = c.value.upper()
        events[c.value] = f"EVT_{evt}" if evt != "ELSE" else "ELSE"

    state_table = states_ws.iter_rows(min_row=2, min_col=6, values_only=True)
    state_table = [list(map(lambda x: x or "", x)) for x in state_table]

    connector_ws: Worksheet = wb["connectors"]
    names = [x[0] for x in (connector_ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True))]
    slot = [x[0] for x in (connector_ws.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True))]
    broadcast = [x[0] for x in (connector_ws.iter_rows(min_row=2, min_col=3, max_col=3, values_only=True))]
    auto_store = [x[0] for x in (connector_ws.iter_rows(min_row=2, min_col=4, max_col=4, values_only=True))]
    enum_name = [x[0] for x in (connector_ws.iter_rows(min_row=2, min_col=5, max_col=5, values_only=True))]
    function_name = [x[0] for x in (connector_ws.iter_rows(min_row=2, min_col=6, max_col=6, values_only=True))]

    connectors = {name: {"slot": s, "broadcast": b, "auto_store": a, "enum_name": e, "funciton_name": f}
                  for name, s, b, a, e, f in zip(names, slot, broadcast, auto_store, enum_name, function_name)}

    config_ws: Worksheet = wb["config"]
    keys = [x[0] for x in config_ws.iter_rows(min_col=1, max_col=1, values_only=True)]
    values = [x[0] for x in config_ws.iter_rows(min_col=2, max_col=2, values_only=True)]
    configs = dict(zip(keys, values))
    return states, events, state_table, connectors, configs


def main(args):
    for f in args.xls:
        wb = read_xls(f)
        states, events, state_table, connectors, configs = parse_workbook(wb)
        header = generate_code(
            configs["class name"],
            states,
            events,
            state_table,
            connectors
        )
        output_path = configs.get("output path", args.output)
        if output_path:
            with open(os.path.join(os.path.dirname(f), output_path), "w") as out:
                out.write(header)
        else:
            print(header)


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument("xls", nargs="+")
    parser.add_argument("--output")
    args = parser.parse_args()
    main(args)
